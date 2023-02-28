import openpyxl
import common.Excel.read_excel.read_excel as read_excel
'''
openpyxl 读取/写入 xlsx/xls 文件的数据时，单元格开始位置为 （1,1）
1、self.book = openpyxl.Workbook()                      创建一个Excel workbook 处理器对象
2、self.sh1 = self.book.create_sheet('年龄表-最前',0)     创建表格  
3、self.sh['A1'] = '你好'
   self.sh.cell(2,2).value = '我爱python'                写入内容
4、book.save                                             保存数据
'''
filename = '../read_excel/work2.xls'
data = read_excel.get_list_excel(filename)
data1 = read_excel.get_dict_excel(filename)

class Write_Xlsx():

    def __init__(self,title,title1):
        '''
        :param title:      第一张表的表名
        :param title1:     第二张表的表名
        '''
        self.book = openpyxl.Workbook()
        # 创建时，会自动产生一个sheet，通过active获取
        self.sh = self.book.active
        # 修改当前 sheet 标题为 工资表
        self.sh.title = title
        # 增加一个 sheet，放在最前
        self.sh1 = self.book.create_sheet(title1,1)

    def write_list_xlsx(self,data,filename):
        '''
        :param data:         需要存储的数据
        :param filename:     需要存储的文件路径
        '''
        # 根据名称获取某个sheet对象  或者直接使用self.sh
        # self.sh = self.book['工资表']
        # 根据行号列号， 写入内容，注意和 xlrd 不同，是从 1 开始
        for i in range(1,len(data)+1):
            for j in range(1,len(data[0])+1):
                self.sh.cell(i,j).value = data[i-1][j-1]
        self.book.save(filename)

    def write_dict_xlsx(self,data,filename):
        '''
        :param data:         需要存储的数据
        :param filename:     需要存储的文件路径
        '''
        # 获取字典的key
        key = [x for x in data[0]]
        for i in range(1,len(data[0])+1):
            self.sh1.cell(1,i).value = key[i-1]
        for j in range(1,len(data)+1):
            for k in range(1,len(key)+1):
                self.sh1.cell(j+1,k).value = data[j-1][key[k-1]]
        self.book.save(filename)

title = '工资表'
title1 = '工资表1'
filename1 = '信息.xlsx'
Write_Xlsx = Write_Xlsx(title,title1)
Write_Xlsx.write_list_xlsx(data,filename1)
Write_Xlsx.write_dict_xlsx(data1,filename1)